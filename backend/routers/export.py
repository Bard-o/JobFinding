"""Export endpoints for CSV and Excel downloads."""

from __future__ import annotations

import csv
import io
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, func
from sqlalchemy.orm import Session

from backend.database import get_db
from backend.models.job import Job
from backend.models.job_technology import JobTechnology
from backend.models.technology import Technology

router = APIRouter(prefix="/export", tags=["export"])


def build_jobs_query(
    db: Session,
    tech: Optional[str] = None,
    seniority: Optional[str] = None,
    work_type: Optional[str] = None,
    country: Optional[str] = None,
    limit: int = 10000,
):
    """Build query with same filters as /jobs endpoint."""
    query = db.query(Job)

    filters = []
    if tech:
        query = query.join(Job.technologies).filter(
            func.lower(Technology.name) == tech.lower()
        )
    if seniority:
        filters.append(Job.seniority == seniority)
    if work_type:
        filters.append(Job.work_type == work_type)
    if country:
        filters.append(Job.country.ilike(f"%{country}%"))

    if filters:
        query = query.filter(and_(*filters))

    return query.limit(limit).all()


def job_to_row(j: Job) -> dict:
    """Convert a Job model instance to a dict for CSV/Excel."""
    return {
        "id": j.id,
        "title": j.title,
        "company": j.company.name if j.company else "",
        "country": j.country or "",
        "published_at": j.published_at.isoformat() if j.published_at else "",
        "url": j.url,
        "work_type": j.work_type or "",
        "seniority": j.seniority or "",
        "description": j.description or "",
        "salary_raw": j.salary_raw or "",
        "technologies": ", ".join(t.name for t in j.technologies),
        "scraped_at": j.scraped_at.isoformat() if j.scraped_at else "",
    }


CSV_HEADERS = [
    "id",
    "title",
    "company",
    "country",
    "published_at",
    "url",
    "work_type",
    "seniority",
    "description",
    "salary_raw",
    "technologies",
    "scraped_at",
]


@router.get("/csv")
def export_csv(
    tech: Optional[str] = Query(default=None),
    seniority: Optional[str] = Query(default=None),
    work_type: Optional[str] = Query(default=None),
    country: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Export jobs matching filters as CSV."""
    jobs = build_jobs_query(
        db, tech=tech, seniority=seniority, work_type=work_type, country=country
    )

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=CSV_HEADERS)
    writer.writeheader()
    for j in jobs:
        writer.writerow(job_to_row(j))

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=jobs_export.csv"},
    )


@router.get("/excel")
def export_excel(
    tech: Optional[str] = Query(default=None),
    seniority: Optional[str] = Query(default=None),
    work_type: Optional[str] = Query(default=None),
    country: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Export jobs matching filters as Excel (.xlsx)."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    jobs = build_jobs_query(
        db, tech=tech, seniority=seniority, work_type=work_type, country=country
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    # Write header row
    for col_idx, header in enumerate(CSV_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data rows
    for row_idx, j in enumerate(jobs, start=2):
        row_data = job_to_row(j)
        for col_idx, header in enumerate(CSV_HEADERS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

    # Auto-size columns
    for col_idx in range(1, len(CSV_HEADERS) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(jobs) + 2)
        )
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=jobs_export.xlsx"},
    )
